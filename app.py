import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import mysql.connector
import time
import logging
from dotenv import load_dotenv
import os

# .env dosyasını yükle
load_dotenv()  # Veritabanı bilgilerini burada güvenli bir şekilde saklayabilirsiniz.

# Veritabanı bağlantısı için sınıf oluşturma
class DatabaseConnection:
    def __init__(self):
        # .env dosyasından alınan bilgiler
        self.host = os.getenv("DB_HOST", "localhost")
        self.user = os.getenv("DB_USER", "root")
        self.password = os.getenv("DB_PASSWORD", "")
        self.database = os.getenv("DB_NAME", "your_database")
        self.conn = None

    def connect(self):
        """Veritabanı bağlantısını başlat."""
        if self.conn is None:
            try:
                self.conn = mysql.connector.connect(
                    host=self.host,
                    user=self.user,
                    password=self.password,
                    database=self.database
                )
            except mysql.connector.Error as err:
                logging.error(f"Veritabanı bağlantısı başarısız: {err}")
                raise

    def get_data(self):
        """Veritabanından veri çek."""
        self.connect()
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM sertifikalar ORDER BY eklenme_tarihi DESC")  # Veritabanını sorgula
        data = cursor.fetchall()  # Veriyi çek
        cursor.close()
        return data

    def close_connection(self):
        """Veritabanı bağlantısını kapat."""
        if self.conn:
            self.conn.close()

# Excel dosyasına veri kaydetme ve sertifika numarasına köprü ekleme
def save_to_excel(data, image_folder):
    excel_folder = 'C:/deneme/ExcelDosyalari'
    os.makedirs(excel_folder, exist_ok=True)
    excel_file = os.path.join(excel_folder, 'Sertifika_Kayitlari.xlsx')

    # Dosya mevcutsa, mevcut dosyadaki verileri silip yeni bir sayfa oluştur
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
        for sheet in workbook.sheetnames:
            std = workbook[sheet]
            workbook.remove(std)

    # Yeni çalışma sayfası oluştur
    workbook = Workbook()

    # Sertifikalar sayfası
    cert_sheet = workbook.active
    cert_sheet.title = "Sertifikalar"
    
    # Başlıkları ekle
    headers = ['Ürün Tanımı', 'Kalite', 'Firma', 'Sertifika No', 'Eklenme Tarihi', 'Sertifika Fotoğrafı']
    cert_sheet.append(headers)

    # Veriyi ekle
    for i, item in enumerate(data, start=2):  # 2'den başlıyoruz, çünkü 1. satır başlık
        added_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        item.append(added_date)  # Eklenme tarihini ekliyoruz

        # Sertifika fotoğrafı yolunu ekle
        cert_image_path = os.path.join(image_folder, f"{item[3]}.jpg")  # Sertifika No'ya göre fotoğraf ismi
        item.append(cert_image_path)

        # Satırda veriyi ekle
        for j, value in enumerate(item[:6], start=1):  # 6 sütun olacak, fotoğraf için ekledik
            cert_sheet.cell(row=i, column=j, value=value)

        # Sertifika No'ya köprü ekle
        cert_sheet[f'D{i}'].hyperlink = cert_image_path
        cert_sheet[f'D{i}'].style = 'Hyperlink'  # Hyperlink stilini uygula

    # Sertifika Fotoğrafı sayfası (resimler)
    img_sheet = workbook.create_sheet(title="Sertifika Fotoğrafı")
    for i, item in enumerate(data, start=1):
        cert_image_path = os.path.join(image_folder, f"{item[3]}.jpg")  # Sertifika numarası ile resim yolu
        if os.path.exists(cert_image_path):
            img = Image(cert_image_path)
            img_sheet.add_image(img, f'A{i}')  # Resmi ekle (A sütununa)

    # Excel dosyasını kaydet
    workbook.save(excel_file)
    return excel_file

# Streamlit arayüzü
st.title('Duyar Metal Kalite Sertifikaları Yönetim Sistemi')

# Veri ekleme kısmı
st.header('Yeni Ürün Ekle')

urun_tanim = st.text_input('Ürün Tanımı:')
kalite = st.text_input('Kalite:')
firma = st.text_input('Firma Adı:')
sertifika_no = st.text_input('Sertifika No:')
sertifika_resmi = st.file_uploader("Sertifika Fotoğrafı Yükle", type=["jpg", "jpeg", "png"])

# Sertifika fotoğrafı yüklenmişse, dosya sistemine kaydedelim
image_folder = 'C:/deneme/SertifikaFotoğrafları'
os.makedirs(image_folder, exist_ok=True)

if sertifika_resmi:
    image_path = os.path.join(image_folder, f"{sertifika_no}.jpg")
    with open(image_path, "wb") as f:
        f.write(sertifika_resmi.getbuffer())

# Yeni ürün eklemek için buton
if st.button('Ürün Ekle'):
    if urun_tanim and kalite and firma and sertifika_no:
        # Yeni veriyi Excel dosyasına ekleyelim
        new_data = [[urun_tanim, kalite, firma, sertifika_no]]  # Sadece gerekli veriler eklendi
        excel_file = save_to_excel(new_data, image_folder)

        # Kullanıcıya Excel dosyasını indirme seçeneği sunalım
        with open(excel_file, "rb") as file:
            st.download_button(
                label="Excel Dosyasını İndir",
                data=file,
                file_name='Sertifika_Kayitlari.xlsx',
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.success('Yeni ürün başarıyla eklendi ve Excel dosyasına kaydedildi.')
    else:
        st.error("Lütfen ürün tanımı, kalite, firma ve sertifika numarası alanlarını doldurun.")

# Arama / Filtreleme Kısmı
st.header('Verileri Ara')

# Excel dosyasını oku
excel_file = 'C:/deneme/ExcelDosyalari/Sertifika_Kayitlari.xlsx'
if os.path.exists(excel_file):
    df = pd.read_excel(excel_file)

    # Filtreleme seçenekleri ekleyelim
    search_term = st.text_input('Arama Yapın (Ürün, Kalite, Firma, Sertifika No)')
    if search_term:
        filtered_df = df[df.apply(lambda row: row.astype(str).str.contains(search_term, case=False).any(), axis=1)]
    else:
        filtered_df = df

    # Filtrelenmiş veriyi göster
    st.dataframe(filtered_df)

else:
    st.warning("Henüz veri eklenmedi.")

# Verileri sürekli güncelleme
st.header('Veri Güncelleme')

# Veritabanındaki verileri güncellemek için buton
if st.button("Veri Güncelle"):
    db_connection = DatabaseConnection()
    try:
        data = db_connection.get_data()
        db_connection.close_connection()
        
        # Veriyi Streamlit üzerinden göster
        st.write(data)  # Burada veriyi tablodan veya istediğiniz formatta gösterebilirsiniz
    except Exception as e:
        st.error(f"Veri güncelleme işlemi başarısız: {e}")
